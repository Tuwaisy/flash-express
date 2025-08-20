"""
Flash Express Bulk Shipments Excel Template Generator
This script creates an Excel template with dropdown menus for bulk shipment creation.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_bulk_shipments_template():
    # Create a new workbook
    wb = Workbook()
    
    # Remove default worksheet and create named sheets
    wb.remove(wb.active)
    
    # Create Shipments sheet
    shipments_sheet = wb.create_sheet("Shipments")
    
    # Create Choices sheet for dropdown data
    choices_sheet = wb.create_sheet("Choices")
    
    # Define the headers for the shipments sheet
    headers = [
        "Client Email",
        "Recipient Name", 
        "Recipient Phone",
        "Package Description",
        "Package Value (EGP)",
        "From Street",
        "From Details", 
        "From City",
        "From Zone",
        "To Street",
        "To Details",
        "To City", 
        "To Zone",
        "Payment Method",
        "Amount to Collect",
        "Is Large Order",
        "Package Weight (kg)",
        "Package Dimensions (LxWxH cm)",
        "Notes"
    ]
    
    # Add headers to shipments sheet
    for col, header in enumerate(headers, 1):
        cell = shipments_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    column_widths = [20, 20, 15, 25, 15, 20, 15, 15, 15, 20, 15, 15, 15, 15, 15, 12, 15, 25, 30]
    for col, width in enumerate(column_widths, 1):
        shipments_sheet.column_dimensions[get_column_letter(col)].width = width
    
    # Define choice data for dropdowns
    choices_data = {
        "Client_Emails": [
            "admin@flash.com",
            "testclient@flash.com", 
            "client@flash.com"
        ],
        "Payment_Methods": [
            "COD",
            "Transfer", 
            "Wallet"
        ],
        "Cities": [
            "Cairo",
            "Alexandria",
            "Giza",
            "Sharm El Sheikh",
            "Hurghada", 
            "Luxor",
            "Aswan",
            "Port Said",
            "Suez",
            "Ismailia"
        ],
        "Cairo_Zones": [
            "Nasr City",
            "Heliopolis",
            "Maadi",
            "Zamalek", 
            "Downtown",
            "Dokki",
            "Mohandessin",
            "New Cairo",
            "6th of October",
            "Shoubra"
        ],
        "Alexandria_Zones": [
            "Downtown",
            "Sidi Gaber",
            "Montaza",
            "Smouha",
            "Miami",
            "Gleem", 
            "Sporting",
            "Stanley",
            "Camp Cesar",
            "Baccos"
        ],
        "Giza_Zones": [
            "Dokki",
            "Mohandessin", 
            "Agouza",
            "Haram",
            "6th of October",
            "Sheikh Zayed",
            "Smart Village",
            "Faisal",
            "Imbaba",
            "Bulaq"
        ],
        "Package_Descriptions": [
            "Electronics - Smartphone",
            "Electronics - Laptop",
            "Electronics - Tablet",
            "Electronics - Headphones",
            "Electronics - Camera",
            "Clothing - T-Shirt",
            "Clothing - Jeans", 
            "Clothing - Shoes",
            "Clothing - Dress",
            "Clothing - Jacket",
            "Books - Novel",
            "Books - Textbook",
            "Books - Magazine",
            "Home & Garden - Kitchen Appliances",
            "Home & Garden - Furniture",
            "Home & Garden - Decor",
            "Food & Beverages - Snacks",
            "Food & Beverages - Drinks",
            "Documents - Legal Papers",
            "Documents - Certificates",
            "Gifts - Birthday Gift",
            "Gifts - Wedding Gift",
            "Medical - Medication",
            "Beauty - Cosmetics",
            "Sports - Equipment"
        ],
        "Package_Values": [
            "50", "100", "150", "200", "250", "300", "400", 
            "500", "750", "1000", "1500", "2000", "3000", "5000"
        ],
        "Large_Order_Options": [
            "TRUE",
            "FALSE"
        ],
        "Street_Examples": [
            "123 Main St",
            "456 Oak Ave", 
            "789 Pine St",
            "321 Cedar Ln",
            "654 Elm Rd",
            "987 Maple Dr",
            "147 Birch Way",
            "258 Willow Ct",
            "369 Palm Blvd",
            "741 Rose Ave"
        ],
        "Building_Details": [
            "Apt 101",
            "Floor 3",
            "Building 5",
            "Suite 201", 
            "Unit 12",
            "Villa 25",
            "Office 304",
            "Shop 15",
            "Penthouse",
            "Ground Floor"
        ]
    }
    
    # Add choice data to choices sheet
    col = 1
    for category, choices in choices_data.items():
        # Add category header
        header_cell = choices_sheet.cell(row=1, column=col, value=category.replace("_", " "))
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center")
        
        # Add choices
        for row, choice in enumerate(choices, 2):
            choices_sheet.cell(row=row, column=col, value=choice)
        
        # Auto-adjust column width
        choices_sheet.column_dimensions[get_column_letter(col)].width = 25
        col += 1
    
    # Create data validation for dropdowns
    def create_dropdown(sheet, column_letter, start_row, end_row, choices_range):
        dv = DataValidation(type="list", formula1=f"Choices.{choices_range}", showDropDown=True)
        dv.error = "Please select a value from the dropdown list"
        dv.errorTitle = "Invalid Entry"
        sheet.add_data_validation(dv)
        dv.add(f"{column_letter}{start_row}:{column_letter}{end_row}")
    
    # Add dropdowns to shipments sheet (assuming 1000 rows for bulk entry)
    end_row = 1000
    
    # Client Email dropdown (Column A)
    create_dropdown(shipments_sheet, "A", 2, end_row, "$A$2:$A$4")
    
    # Payment Method dropdown (Column N) 
    create_dropdown(shipments_sheet, "N", 2, end_row, "$B$2:$B$4")
    
    # Cities dropdown for From City (Column H) and To City (Column L)
    create_dropdown(shipments_sheet, "H", 2, end_row, "$C$2:$C$11")
    create_dropdown(shipments_sheet, "L", 2, end_row, "$C$2:$C$11")
    
    # Package Description dropdown (Column D)
    create_dropdown(shipments_sheet, "D", 2, end_row, "$G$2:$G$26")
    
    # Package Value dropdown (Column E)
    create_dropdown(shipments_sheet, "E", 2, end_row, "$H$2:$H$15")
    
    # Large Order dropdown (Column P)
    create_dropdown(shipments_sheet, "P", 2, end_row, "$I$2:$I$3")
    
    # Add sample data rows
    sample_data = [
        [
            "testclient@flash.com", "John Doe", "+201234567890", "Electronics - Smartphone", 
            "500", "123 Main St", "Apt 101", "Cairo", "Nasr City", 
            "456 Oak Ave", "Building 5", "Alexandria", "Downtown", "COD", 
            "520", "FALSE", "0.5", "15x10x5", "Sample shipment 1"
        ],
        [
            "testclient@flash.com", "Jane Smith", "+201987654321", "Clothing - T-Shirt",
            "150", "123 Main St", "Apt 101", "Cairo", "Nasr City",
            "789 Pine St", "Floor 3", "Giza", "Dokki", "Transfer",
            "0", "FALSE", "0.3", "20x15x2", "Sample shipment 2"
        ]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            shipments_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Add instructions sheet
    instructions_sheet = wb.create_sheet("Instructions")
    instructions_sheet.column_dimensions["A"].width = 80
    
    instructions = [
        "FLASH EXPRESS BULK SHIPMENTS TEMPLATE",
        "",
        "INSTRUCTIONS:",
        "1. Use the 'Shipments' sheet to enter your bulk shipment data",
        "2. Most columns have dropdown menus - click the arrow to select from predefined options",
        "3. The 'Choices' sheet contains all the dropdown options - you can modify these as needed",
        "",
        "REQUIRED FIELDS:",
        "• Client Email - Must be a valid registered client email",
        "• Recipient Name - Full name of the person receiving the package", 
        "• Recipient Phone - Phone number in format +201234567890",
        "• Package Description - Brief description of package contents",
        "• Package Value - Monetary value of the package in EGP",
        "• From/To Address Fields - Complete address information",
        "• Payment Method - COD, Transfer, or Wallet",
        "",
        "NOTES:",
        "• For COD shipments, 'Amount to Collect' = Package Value + Shipping Fee",
        "• For Transfer shipments, 'Amount to Collect' can be 0 if no additional amount needed",
        "• For Wallet shipments, 'Amount to Collect' should typically be 0",
        "• Is Large Order: TRUE for packages requiring special handling",
        "• Package Weight in kilograms (e.g., 0.5, 1.2, 2.0)",
        "• Package Dimensions in LxWxH format in centimeters (e.g., 15x10x5)",
        "",
        "AFTER COMPLETING:",
        "1. Save this file",
        "2. Upload it to the Flash Express bulk shipment import feature",
        "3. Review the preview before confirming the import",
        "",
        "For support, contact: admin@flash.com"
    ]
    
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=row, column=1, value=instruction)
        if row == 1:
            cell.font = Font(bold=True, size=16, color="366092")
        elif instruction.endswith(":") and not instruction.startswith("•"):
            cell.font = Font(bold=True, color="70AD47")
        cell.alignment = Alignment(wrap_text=True)
    
    # Set the shipments sheet as the active sheet
    wb.active = shipments_sheet
    
    # Save the workbook
    filename = "Flash_Express_Bulk_Shipments_Template.xlsx"
    wb.save(filename)
    print(f"✅ Template created successfully: {filename}")
    print(f"📋 The template includes:")
    print(f"   • Shipments sheet with dropdown menus")
    print(f"   • Choices sheet with all dropdown options") 
    print(f"   • Instructions sheet with detailed guidance")
    print(f"   • Sample data rows for reference")

if __name__ == "__main__":
    try:
        create_bulk_shipments_template()
    except ImportError as e:
        print("❌ Missing required libraries. Please install them with:")
        print("pip install pandas openpyxl")
        print(f"Error: {e}")
    except Exception as e:
        print(f"❌ Error creating template: {e}")
